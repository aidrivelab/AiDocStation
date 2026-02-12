# -*- coding: utf-8 -*-
"""
@File    : local_bridge/service/spreadsheet/generator.py
@Desc    : AiDoc Station Lite 核心模块 - 赋能高效文档协作与智能排版处�?
@Author  : AIDriveLab Team
@Create  : 2026-02-09 21:12:43
@Version : V0.2.6
@Copyright: ©AIDriveLab Inc. All Rights Reserved.
"""

from typing import List
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

from ...utils.logging import log
from ...core.errors import InsertError
from .formatting import CellFormat


class SpreadsheetGenerator:
    




       
    
    @staticmethod
    def generate_xlsx_bytes(table_data: List[List[str]], keep_format: bool = True) -> bytes:
        











           
        try:

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            if not table_data:
                log("Table data is empty, creating empty spreadsheet")

                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                return buffer.read()
            

            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            code_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            

            for row_idx, row_data in enumerate(table_data, start=1):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    if keep_format:

                        cell_format = CellFormat(cell_value)
                        clean_text = cell_format.parse()
                        

                        hyperlink_url = None
                        if cell_format.segments:

                            for seg in cell_format.segments:
                                if seg.hyperlink_url:
                                    hyperlink_url = seg.hyperlink_url
                                    break
                        

                        if cell_format.has_newline:
                            cell.alignment = Alignment(wrap_text=True, vertical="top")
                        
                        if cell_format.is_code_block:

                            cell.value = clean_text
                            cell.font = Font(name="Consolas")
                            cell.fill = code_fill
                            cell.alignment = Alignment(wrap_text=True, vertical="top")
                        elif hyperlink_url:

                            cell.value = clean_text
                            cell.hyperlink = hyperlink_url
                            cell.font = Font(color="0563C1", underline="single")
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        elif len(cell_format.segments) > 1:

                            rich_text_parts = []
                            has_inline_code = False
                            
                            for seg in cell_format.segments:
                                if not seg.text:
                                    continue
                                

                                if seg.is_code:
                                    has_inline_code = True
                                

                                inline_font = InlineFont(
                                    b=seg.bold,
                                    i=seg.italic,
                                    strike=seg.strikethrough,
                                    rFont="Consolas" if seg.is_code else None
                                )
                                

                                rich_text_parts.append(TextBlock(inline_font, seg.text))
                            

                            if rich_text_parts:
                                cell.value = CellRichText(*rich_text_parts)
                                

                                if has_inline_code:
                                    cell.fill = code_fill
                        elif len(cell_format.segments) == 1:

                            seg = cell_format.segments[0]
                            cell.value = clean_text
                            

                            has_inline_code = seg.is_code
                            if has_inline_code:
                                cell.fill = code_fill
                            

                            if seg.bold or seg.italic or seg.strikethrough or seg.is_code:
                                cell.font = Font(
                                    bold=seg.bold,
                                    italic=seg.italic,
                                    strike=seg.strikethrough,
                                    name="Consolas" if seg.is_code else None
                                )
                        else:

                            cell.value = clean_text
                    else:

                        cell_format = CellFormat(cell_value)
                        cell.value = cell_format.parse()
                    

                    if row_idx == 1:
                        cell.fill = header_fill
                        cell.font = Font(bold=True)
                    

                    if not cell.alignment or not cell.alignment.wrap_text:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            

            for col_idx in range(1, len(table_data[0]) + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                
                for row_idx in range(1, len(table_data) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    try:
                        if cell.value:

                            lines = str(cell.value).split('\n')
                            max_line_length = max(len(line) for line in lines) if lines else 0
                            if max_line_length > max_length:
                                max_length = max_line_length
                    except Exception:
                        pass
                

                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[col_letter].width = adjusted_width
            

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            xlsx_bytes = buffer.read()
            log(f"Successfully generated XLSX bytes: {len(xlsx_bytes)} bytes")
            return xlsx_bytes
            
        except Exception as e:
            log(f"Failed to generate XLSX: {e}")
            raise InsertError(f"生成 XLSX 文件失败: {e}")
